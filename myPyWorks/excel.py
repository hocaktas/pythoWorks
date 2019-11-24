# bu program bir excel dosyasının istenilen bölümünden okuduğu verileri
#başka bir excel dosyasının istediğin kısmına kaydeder
#geliştirilebilir

from openpyxl import *

kitap=load_workbook("IOK10006R1_621.xlsx")
sayfa=kitap.active
ogrno=[]
ogradi=[]
notu=[]

for i in range(16,795):
    if sayfa.cell(i,2).value!=None:
        ogrno.append(sayfa.cell(i,2).value)
        ogradi.append(sayfa.cell(i,3).value)
        notu.append(sayfa.cell(i,18).value)
print(len(ogradi))

kitap2=Workbook()

sheet = kitap2.active
k=0
for i in range(k,len(ogrno)):
    sheet.cell(k+1, 1, ogrno[k])
    sheet.cell(k+1, 2, ogradi[k])
    sheet.cell(k+1, 3, notu[k])
    k+=1

kitap2.save("C:/Users/aktas/OneDrive/Masaüstü/calisma.xlsx")


kitap2.close()
kitap.close()